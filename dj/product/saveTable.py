import time

import requests
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
from .trans_rmb import money_en_to_cn
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side

border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

nameMap = {
    "name": '名称',
    "specification": '规格',
    "kind": '类型',
    "level": '级别',
    "unit": '单位',
    "weight": '单重',
    "quantity": '数量',
    "price": '价格/kg',
    "unitprice": '单价',
    "amount": '合计'
}

data = [
    {
        "name": "插堵",
        "specification": "80",
        "kind": "K型",
        "level": "None",
        "unit": "件",
        "weight": "6.50",
        "quantity": "12321",
        "price": "323",
        "rowid": "row_31",
        "unitprice": "2099.50",
        "amount": "25867939.50"
    },
    {
        "name": "螺栓",
        "specification": "100～500",
        "kind": "K型",
        "level": "None",
        "unit": "米",
        "weight": "0.40",
        "quantity": "1",
        "price": "32",
        "rowid": "row_35",
        "unitprice": "12.80",
        "amount": "12.80"
    }
]

op = {
    "name": True,
    "specification": True,
    "kind": True,
    "level": True,
    "unit": True,
    "weight": True,
    "quantity": True,
    "price": True,
    "unitprice": True,
    "amount": True
}
title_list = ["name",
              "specification",
              "kind",
              "level",
              "unit",
              "weight",
              "quantity",
              "price",
              "unitprice",
              "amount"]


def f_name(s):
    if s is None:
        return ""
    elif s == "None":
        return ""
    else:
        return s


def export_to_excel(table, options):
    # 创建一个新的工作簿

    wb = Workbook()
    ws = wb.active

    # 写入表头
    header_labels = ["序号"]
    for item in title_list:
        if options[item]:
            header_labels.append(nameMap[item])
    header_labels.append("备注")
    max_col = len(header_labels)
    max_row = len(table)
    crow = 1

    ws.merge_cells(None, crow, 1, crow, max_col)
    ws.cell(row=crow, column=1, value="致：").alignment = Alignment(horizontal='left')

    crow += 1

    ws.merge_cells(None, crow, 1, crow, max_col)
    ws.cell(row=crow, column=1, value="联系人：        电话：").alignment = Alignment(horizontal='left')
    crow += 1

    ws.merge_cells(None, crow, 1, crow, max_col)
    ws.cell(row=crow, column=1, value="安钢集团永通球墨铸铁管报价单").alignment = Alignment(horizontal='center')
    crow += 1

    for col, label in enumerate(header_labels, start=1):
        ws.cell(row=crow, column=col, value=label).alignment = Alignment(horizontal='center')
    crow += 1

    # 写入表格数据
    # ws.cell(row=1, column=1, value=f"序号").alignment = Alignment(horizontal='center')

    ans = 0.0
    for row in range(max_row):
        col = 1
        ws.cell(row=crow, column=col, value=f"{row + 1}").alignment = Alignment(horizontal='center')
        col += 1

        for item in title_list:
            if options[item]:
                ws.cell(row=crow, column=col, value=f"{f_name(table[row][item])}").alignment = Alignment(
                    horizontal='center')
                col += 1
        ws.cell(row=crow, column=col, value=f"").alignment = Alignment(horizontal='center')
        ans += float(table[row]['amount'])
        crow += 1
    ans = round(ans, 2)
    # ws.cell(row=crow, column=1, value=f"{crow-1}").alignment = Alignment(horizontal='center')
    # ws.cell(row=crow, column=2, value=f"合计").alignment = Alignment(horizontal='center')

    ws.merge_cells(None, crow, 3, crow, max_col - 2)
    ws.cell(row=crow, column=1, value=f"{max_row + 1}").alignment = Alignment(horizontal='center')
    ws.cell(row=crow, column=2, value=f"合计").alignment = Alignment(horizontal='center')
    ws.cell(row=crow, column=3, value=f"{money_en_to_cn(ans)}").alignment = Alignment(horizontal='center')
    ws.cell(row=crow, column=max_col - 1, value=f"{ans}").alignment = Alignment(horizontal='center')
    crow += 1

    # ws.merge_cells(f'A{crow}:A{crow + 1}')
    ws.cell(row=crow, column=1, value=f"说明").alignment = Alignment(horizontal='center')

    ws.merge_cells(None, crow, 2, crow, max_col)
    # ws.cell(row=crow, column=2, value="西安永通球墨铸铁管营销有限公司").alignment = Alignment(
    #     horizontal='right')
    crow += 1
    ws.merge_cells(None, crow, 1, crow, max_col)
    ws.cell(row=crow, column=1, value="西安永通球墨铸铁管营销有限公司").alignment = Alignment(
        horizontal='right')
    crow += 1

    ws.merge_cells(None, crow, 1, crow, max_col)
    ws.cell(row=crow, column=1, value="联系人：  电话： 传真：0298569781").alignment = Alignment(
        horizontal='right')
    crow += 1

    ws.merge_cells(None, crow, 1, crow, max_col)
    ws.cell(row=crow, column=1, value=f"{time.strftime('%Y年%m月%d日')}").alignment = Alignment(
        horizontal='right')
    crow += 1
    for i in range(3, crow):
        for j in range(1, max_col + 1):
            ws.cell(row=i, column=j).border = border
    # for col_idx, col_width in enumerate([10, 15, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, ]):
    #     ws.column_dimensions[chr(65 + col_idx)].width = col_width
    for i in range(1, max_col):
        ws.column_dimensions[get_column_letter(i)].width = 13

    # 保存工作簿为 Excel 文件
    loca = time.strftime('./tables/%Y-%m-%d-%H-%M-%S')
    # loca = "test"
    excel_filename = f'{loca}.xlsx'
    wb.save(excel_filename)
    return excel_filename
    # print(f"表格数据已成功导出到 {excel_filename}")

# export_to_excel(data, op)
